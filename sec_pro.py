import openpyxl

# 打开文本文件
with open('uORF_final_list.txt', 'r') as file:
    lines = file.readlines()

# 创建一个新的Excel工作簿
wb = openpyxl.Workbook()
sheet = wb.active

# 添加表头
sheet['A1'] = '文件名'
sheet['B1'] = '二级结构成分'
sheet['C1'] = '统计值'

# 遍历文本文件的每一行
for row, line in enumerate(lines, start=2):
    # 分割文件名和二级结构成分
    parts = line.split(' ', 1)
    if len(parts) == 2:
        file_name, secondary_structure = parts
        file_name = file_name.strip()
        secondary_structure = secondary_structure.strip()

        # 统计二级结构成分中特定字符的总数
        total_chars = len(secondary_structure.replace(" ", ""))
        count_H = secondary_structure.count('H')
        count_G = secondary_structure.count('G')
        count_I = secondary_structure.count('I')
        count_E = secondary_structure.count('E')
        count_B = secondary_structure.count('B')
        count_plus = secondary_structure.count('+')

        # 计算统计值
        if total_chars != 0:
            ratio = (count_H + count_G + count_I + count_E + count_B + count_plus) / total_chars 
        else:
            ratio = 0

        # 将数据写入Excel表格
        sheet.cell(row=row, column=1, value=file_name)
        sheet.cell(row=row, column=2, value=secondary_structure)
        sheet.cell(row=row, column=3, value=ratio)

# 保存Excel文件
wb.save('sec_431.xlsx')
