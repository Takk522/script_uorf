import openpyxl

# 打开Excel文件
workbook = openpyxl.load_workbook('dssp_1428.xlsx')
sheet = workbook.active

# 添加表头（如果需要）
sheet['A1'] = 'file_name'
sheet['B1'] = 'secondary_structure'
sheet['C1'] = 'secondary_structure_component'

# 遍历Excel文件的每一行，跳过第一行（表头）
for row, row_data in enumerate(sheet.iter_rows(min_row=2), start=2):
    # 从Excel表格中获取文件名和二级结构成分
    file_name = row_data[0].value
    secondary_structure = row_data[1].value

    # 统计二级结构成分中特定字符的总数
    total_chars = len(secondary_structure)
    count_H = secondary_structure.count('H')
    count_G = secondary_structure.count('G')
    count_I = secondary_structure.count('I')
    count_E = secondary_structure.count('E')
    count_B = secondary_structure.count('B')

    if total_chars != 0:
        ratio = (count_H + count_G + count_I + count_E + count_B) / total_chars
    else:
        ratio = 0

    sheet.cell(row=row, column=3, value=ratio)

# 保存修改后的Excel文件
workbook.save('sec_1428.xlsx')
